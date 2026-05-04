## @file exporter.py
#  @brief Servicio para la generación de reportes en formato Excel (.xlsx).

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    def __init__(self, torneo_repo, prueba_repo, participacion_repo):
        self.torneo_repo = torneo_repo
        self.prueba_repo = prueba_repo
        self.participacion_repo = participacion_repo

    def generar_reporte_torneo(self, id_torneo, path_archivo):
        """Genera el reporte agrupado por instancias y con columna Club."""
        torneo = self.torneo_repo.obtener_por_id(id_torneo)
        pruebas = self.prueba_repo.obtener_por_torneo(id_torneo)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Estilos
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        sub_header_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        row = 1
        # --- ENCABEZADO DEL TORNEO (Nombre - Fecha) ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell_t = ws.cell(row=row, column=1, value=f"TORNEO: {torneo.nombre.upper()} - {torneo.fecha_inicio}")
        cell_t.fill = header_fill
        cell_t.font = white_font
        cell_t.alignment = Alignment(horizontal="center")
        row += 2

        for prueba in pruebas:
            sexo_str = "MASCULINO" if prueba.sexo == "M" else "FEMENINO"
            base_info = f"PRUEBA: {prueba.nombre} - {prueba.categoria} - {sexo_str}"
            
            # Traer resultados
            participantes = self.participacion_repo.obtener_resultados_prueba(prueba.id_prueba)
            
            # Agrupar por instancia (Serie 1, Final, etc.)
            grupos = {}
            for p in participantes:
                if p.instancia not in grupos: grupos[p.instancia] = []
                grupos[p.instancia].append(p)

            for instancia, atletas in grupos.items():
                # Título de la instancia
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                ws.cell(row=row, column=1, value=f"{base_info} - {instancia.upper()}").font = sub_header_font
                row += 1

                # Encabezados de tabla
                headers = ["Pos", "Nombre", "Apellido", "Club", "Tiempo"]
                for c, text in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=c, value=text)
                    cell.font = Font(bold=True)
                    cell.border = border
                row += 1

                # Datos
                for i, p in enumerate(atletas, 1):
                    pos = f"{i}" if p.resultado and p.resultado != "---" else "-"
                    # Si no tiene club, ponemos "Libre"
                    club = getattr(p, 'club', "---")
                    
                    datos = [pos, p.nombre, p.apellido, club, p.resultado if p.resultado else "---"]
                    for c, val in enumerate(datos, 1):
                        cell = ws.cell(row=row, column=c, value=val)
                        cell.border = border
                    row += 1
                row += 1 # Espacio entre tablas

        # Ajuste de ancho
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(path_archivo)